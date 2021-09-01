# open 函数
# 文件写入
with open('./data.txt', mode='a', encoding='utf-8') as f:
    f.write("summer\n")
    f.writelines(['junjun\n', 'yueyue\n'])
f.write('秋天到了')

# 文件读取
with open('./data.txt', mode='r', encoding='utf-8') as f:
    # lines = f.read()  # 读取文件所有内容
    # print(lines)
    lines = f.readlines()  # 返回list数据
    print(lines)

# csv 操作
# 写入文件
import csv

with open('./data.csv', mode='w', newline='', encoding='utf-8')as file:
    fw = csv.writer(file)
    fw.writerow(['tianqi', 'yintian'])
    fw.writerow(['yuefen', 'jiuyue'])
    fw.writerows([('he', 'yangyang'), ('she', 'yueyue')])

# dict 方式写入数据
import csv

with open('./data1.csv', mode='w', newline='', encoding='utf-8')as file:
    filed_name = ['user', 'passwd']
    cdw = csv.DictWriter(file, fieldnames=filed_name)  # 定制表头
    cdw.writeheader()  # 写入表头
    cdw.writerow({'user': 'he', 'passwd': '12345'})
    cdw.writerow({'user': 'she', 'passwd': '67890'})
    cdw.writerows([{'user': 'he', 'passwd': '12345'}, {'user': 'she', 'passwd': '67890'}])
# 读取文件
import csv

with open('./data.csv', mode='r', newline='', encoding='utf-8')as file:
    cr = csv.reader(file)
    for line in cr:
        print(line)

# excel 操作
# 写入文件
from openpyxl import Workbook

wb = Workbook()
worksheet = wb.create_sheet('data')
worksheet.append(['username', 'password'])
for i in range(100):
    worksheet.append([f'testuser{i}', '123456'])

wb.save('./data.xlsx')

# 读取文件
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

wb = load_workbook('data.xlsx')
print(wb.sheetnames)
ws: Worksheet = wb['data']
test_data = []
for row in ws.values:
    test_data.append(row)

print(test_data)

# json
# 转换为字符串
import json

jstr = json.dumps({'he': 'yangyang'})
print(jstr, type(jstr))
with open('./data.txt', mode='w')as file:
    file.write(jstr)
# json解析为python 字符串
import json

ls = json.loads('["hello", {"username": "xiaoming"}]')
print(ls, type(ls))

# 写入json文件
import json

test_data = {
    "id": 1,
    "test_steps": [
        "1.打开浏览器",
        "2.输入用户信息"
    ]
}

with open('./data.json', mode='w', encoding='utf-8') as f:
    json.dump(test_data, f)

# 读取json

import json

with open('./data.json', mode='r', encoding='utf-8') as f:
    data = json.load(f)
    print(data, type(data))

# yaml
# 解析数据
import yaml

data = """
# - Cat
# - Dog
# - Goldfish
"""
print(yaml.safe_load(data), type(yaml.safe_load(data)))  # list

data1 = {
    "name": 'xiaoming',
    'age': 20
}

print(yaml.safe_dump(data1), type(yaml.safe_dump(data1)))  # str

# 写入文件
import yaml

data = {
    'data': [1, 2, 3, 4, 5],
    "name": 'xiaoming',
    'age': 20
}

with open('./data.yml', mode='w', encoding='utf-8') as f:
    yaml.safe_dump(data, f)

import yaml

with open('./data.yml', mode='r', encoding='utf-8') as f:
    print(yaml.safe_load(f))

# ini 文件解析
# 写入ini文件
import configparser

# 实例化对象
config = configparser.ConfigParser()

config['Default'] = {
    "IP": "127.0.0.1",
    "port": 3000,
    "user": "root",
    "passwd": 123456
}

config['uat'] = {
    "IP": "192.168.1.101",
    "port": 3000,
    "user": "root",
    "passwd": 123456
}
with open('./config.ini', mode='w', encoding='utf-8')as f:
    config.write(f)

# 读取ini文件
import configparser

# 实例化对象
config = configparser.ConfigParser()
config.read('./config.ini')
print(config.sections())
test_data = {}
for section in config.sections():
    # print(section, config[section])
    test_data[section] = {}
    for key in config[section]:
        test_data[section][key] = config[section][key]
        # print(key, config[section].get(key))
print(test_data)
